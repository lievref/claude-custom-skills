#!/usr/bin/env python
"""Flatten the DNIT ICM sheet and export directly to xlsx/csv/gpkg."""

from __future__ import annotations

import argparse
import csv
import math
import sys
from pathlib import Path

import geopandas as gpd
from openpyxl import Workbook, load_workbook
from shapely import wkt as shapely_wkt
from shapely.geometry import LineString, MultiLineString
from shapely.ops import linemerge


SOURCE_SHEET = "ICM"
FLAT_SHEET = "Export_ICM"
DATA_START_ROW = 9
MAPPINGS = [
    ("A", "Contrato"),
    ("B", "Codigo_SNV"),
    ("C", "UF"),
    ("D", "BR"),
    ("E", "km_inic"),
    ("F", "km_fina"),
    ("G", "Sentido"),
    ("H", "Superficie"),
    ("I", "ICP-Panela"),
    ("M", "ICP-Remendo"),
    ("Q", "ICP-Trincamento"),
    ("U", "ICC-Roçada"),
    ("X", "ICC-Drenagem"),
    ("AB", "ICC-Sinalização"),
    ("AI", "ICP"),
    ("AJ", "ICC"),
    ("AK", "ICM"),
]


def parse_decimal(value: object) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".") if text.count(",") == 1 and text.count(".") > 1 else text
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def prepare_linestring(geometry):
    if geometry is None or geometry.is_empty:
        return None
    if isinstance(geometry, LineString):
        return geometry
    if isinstance(geometry, MultiLineString):
        merged = linemerge(geometry)
        if isinstance(merged, LineString):
            return merged
        if hasattr(merged, "geoms"):
            return max(list(merged.geoms), key=lambda geom: geom.length)
    return geometry


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Flatten ICM and export directly.")
    parser.add_argument("input_workbook", help="Path to the source workbook")
    parser.add_argument(
        "--formats",
        default="gpkg",
        help="Comma-separated output formats: gpkg,csv,xlsx",
    )
    parser.add_argument(
        "--output-base",
        default=None,
        help="Output path without extension. Default: <input>_export_icm",
    )
    parser.add_argument(
        "--source-sheet",
        default=SOURCE_SHEET,
        help=f"Source sheet name. Default: {SOURCE_SHEET}",
    )
    parser.add_argument(
        "--flat-sheet",
        default=FLAT_SHEET,
        help=f"Flattened sheet name used for geometry lookup. Default: {FLAT_SHEET}",
    )
    return parser.parse_args()


def load_workbook_rows(path: Path, source_sheet: str, flat_sheet: str) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[source_sheet]
    records: list[dict] = []

    geometry_lookup = {}
    if flat_sheet in workbook.sheetnames:
        flat = workbook[flat_sheet]
        flat_headers = [flat.cell(1, c).value for c in range(1, flat.max_column + 1)]
        if "geometry_wkt" in flat_headers:
            flat_idx = {str(flat_headers[i - 1]): i for i in range(1, len(flat_headers) + 1) if flat_headers[i - 1]}
            for row_idx in range(2, flat.max_row + 1):
                key = (
                    flat.cell(row_idx, flat_idx["BR"]).value,
                    flat.cell(row_idx, flat_idx["km_inic"]).value,
                    flat.cell(row_idx, flat_idx["km_fina"]).value,
                    flat.cell(row_idx, flat_idx["Sentido"]).value,
                )
                geometry_lookup[key] = flat.cell(row_idx, flat_idx["geometry_wkt"]).value

    for row_idx in range(DATA_START_ROW, worksheet.max_row + 1):
        if not str(worksheet[f"D{row_idx}"].value or "").strip():
            continue
        record = {target: worksheet[f"{source}{row_idx}"].value for source, target in MAPPINGS}
        geo_key = (record["BR"], record["km_inic"], record["km_fina"], record["Sentido"])
        if geo_key in geometry_lookup:
            record["geometry_wkt"] = geometry_lookup[geo_key]
        records.append(record)

    return records


def write_xlsx(path: Path, rows: list[dict], sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    headers = list(rows[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(1, col_idx, header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row_idx, col_idx, row.get(header))
    workbook.save(path)


def write_csv(path: Path, rows: list[dict]) -> None:
    headers = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def build_km_markers(spatial_rows: list[dict]) -> list[dict]:
    groups: dict[tuple[str, str], list[dict]] = {}
    for row in spatial_rows:
        geometry = prepare_linestring(row.get("geometry"))
        km_ini = parse_decimal(row.get("km_inic"))
        km_fim = parse_decimal(row.get("km_fina"))
        if geometry is None or km_ini is None or km_fim is None or math.isclose(km_ini, km_fim):
            continue
        key = (str(row.get("BR") or "").strip(), str(row.get("UF") or "").strip())
        groups.setdefault(key, []).append(
            {
                "geometry": geometry,
                "km_ini": km_ini,
                "km_fim": km_fim,
                "km_min": min(km_ini, km_fim),
                "km_max": max(km_ini, km_fim),
            }
        )

    marker_rows: list[dict] = []
    tol = 1e-9
    for (br, uf), segments in groups.items():
        intervals = sorted((segment["km_min"], segment["km_max"]) for segment in segments)
        merged: list[list[float]] = []
        for start, end in intervals:
            if not merged or start > merged[-1][1] + tol:
                merged.append([start, end])
            else:
                merged[-1][1] = max(merged[-1][1], end)

        for start, end in merged:
            km_start = math.floor(start + tol) + 1
            km_end = math.ceil(end - tol) - 1
            for km in range(km_start, km_end + 1):
                point = None
                for segment in segments:
                    if segment["km_min"] - tol <= km <= segment["km_max"] + tol:
                        fraction = (km - segment["km_ini"]) / (segment["km_fim"] - segment["km_ini"])
                        fraction = max(0.0, min(1.0, fraction))
                        point = segment["geometry"].interpolate(fraction, normalized=True)
                        break
                if point is None:
                    continue
                marker_rows.append(
                    {
                        "BR": br,
                        "UF": uf,
                        "km": km,
                        "rotulo": f"km {km}",
                        "geometry": point,
                    }
                )
    return marker_rows


def write_gpkg(path: Path, rows: list[dict]) -> None:
    if "geometry_wkt" not in rows[0]:
        raise RuntimeError("No geometry_wkt column available. Generate geometry before exporting to GeoPackage.")

    spatial_rows = []
    for row in rows:
        new_row = dict(row)
        geom_text = new_row.pop("geometry_wkt", None)
        new_row["geometry_wkt"] = geom_text
        new_row["geometry"] = shapely_wkt.loads(geom_text) if geom_text else None
        spatial_rows.append(new_row)

    if path.exists():
        path.unlink()

    gdf = gpd.GeoDataFrame(spatial_rows, geometry="geometry", crs="EPSG:4674")
    layer = path.stem[:63]
    gdf.to_file(path, layer=layer, driver="GPKG")

    marker_rows = build_km_markers(spatial_rows)
    if marker_rows:
        markers_gdf = gpd.GeoDataFrame(marker_rows, geometry="geometry", crs="EPSG:4674")
        markers_gdf.to_file(path, layer="marcos-km", driver="GPKG")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_workbook)
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    output_base = Path(args.output_base) if args.output_base else input_path.with_name(f"{input_path.stem}_export_icm")
    rows = load_workbook_rows(input_path, args.source_sheet, args.flat_sheet)
    if not rows:
        raise RuntimeError("No rows extracted from the source sheet.")

    formats = [fmt.strip().lower() for fmt in args.formats.split(",") if fmt.strip()]
    for fmt in formats:
        if fmt == "xlsx":
            write_xlsx(output_base.with_suffix(".xlsx"), rows, args.flat_sheet)
        elif fmt == "csv":
            write_csv(output_base.with_suffix(".csv"), rows)
        elif fmt == "gpkg":
            write_gpkg(output_base.with_suffix(".gpkg"), rows)
        else:
            raise RuntimeError(f"Unsupported format: {fmt}")

    for fmt in formats:
        print(output_base.with_suffix(f".{fmt}"))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        raise SystemExit(1)
