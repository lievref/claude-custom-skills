#!/usr/bin/env python
"""SNV helpers for point lookup, segment extraction, and table enrichment."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import unicodedata
from pathlib import Path


EPSG_GEO = 4674
EPSG_UTM = 31983  # SIRGAS 2000 / UTM zone 23S — used only as last-resort fallback
FUSOS_UTM_PATH = "/home/liev/.claude/fusos-utm-brasil/FUSOS_UTM.shp"
DEFAULT_TABLE_SHEET = "Export_ICM"
TABLE_SHEET_ALIASES = ("Export_ICM", "Export-ICM")

SNV_FIELD_MAP = {
    "vl_br": ["vl_br", "br", "vlbr"],
    "sg_uf": ["sg_uf", "uf", "sguf"],
    "vl_km_inic": ["vl_km_inic", "km_inic", "vlkminic", "km_inicial"],
    "vl_km_fina": ["vl_km_fina", "km_fina", "vlkmfina", "km_final"],
    "vl_extensa": ["vl_extensa", "vl_ext", "extensa", "extensao", "vl_extensao"],
    "vl_codigo": ["vl_codigo", "codigo", "cod_trecho", "id_trecho"],
    "versao_snv": ["versao_snv", "versao", "versao_base"],
}

OPTIONAL_SNV_FIELDS = {"vl_codigo", "versao_snv", "vl_extensa"}

TABLE_FIELD_ALIASES = {
    "br": ["br", "rodovia", "rod", "highway", "br_uf"],
    "uf": ["uf", "estado", "sg_uf", "unidade_federativa"],
    "kmi": ["kmi", "km inicial", "km_ini", "km_inic", "km_inicial", "quilometro inicial", "km de"],
    "kmf": ["kmf", "km final", "km_fim", "km_fina", "km_final", "quilometro final", "km ate"],
}


def _normalize_name(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def _clean_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def _to_float(value: object) -> float | None:
    text = _clean_value(value)
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".") if text.count(",") == 1 and text.count(".") > 1 else text
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_br(value: object) -> str:
    text = _clean_value(value).upper()
    match = re.search(r"(\d+)", text)
    if not match:
        return ""
    return match.group(1).lstrip("0") or "0"


def _normalize_uf(value: object, default: str | None = None) -> str:
    text = _clean_value(value).upper()
    if text:
        return text
    return (default or "").upper()


def _normalize_sheet_name(value: object) -> str:
    return _normalize_name(value)


def _pick_sheet_name(sheet_names, requested: str | None = None) -> str:
    normalized_to_real = {_normalize_sheet_name(name): name for name in sheet_names}
    if requested:
        target = normalized_to_real.get(_normalize_sheet_name(requested))
        if target:
            return target
        raise RuntimeError(
            f"Worksheet '{requested}' not found. Available sheets: {list(sheet_names)}"
        )

    for alias in TABLE_SHEET_ALIASES:
        target = normalized_to_real.get(_normalize_sheet_name(alias))
        if target:
            return target

    raise RuntimeError(
        f"Expected worksheet '{DEFAULT_TABLE_SHEET}' not found. Available sheets: {list(sheet_names)}"
    )


def _trim_headers(values):
    headers = [_clean_value(value) for value in values]
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def _resolve_snv_fields(columns) -> dict[str, str | None]:
    normalized = {_normalize_name(col): col for col in columns}
    resolved = {}
    for key, aliases in SNV_FIELD_MAP.items():
        found = None
        for alias in aliases:
            found = normalized.get(_normalize_name(alias))
            if found is not None:
                break
        if found is None and key not in OPTIONAL_SNV_FIELDS:
            raise KeyError(
                f"Required SNV field '{key}' not found. Available columns: {list(columns)}"
            )
        resolved[key] = found
    return resolved


def _resolve_table_fields(headers) -> dict[str, str | None]:
    normalized = {_normalize_name(col): col for col in headers}
    resolved = {}
    for key, aliases in TABLE_FIELD_ALIASES.items():
        found = None
        for alias in aliases:
            found = normalized.get(_normalize_name(alias))
            if found is not None:
                break
        resolved[key] = found
    if resolved["br"] is None or resolved["kmi"] is None or resolved["kmf"] is None:
        raise KeyError(
            "Input table must contain aliases for BR, kmi, and kmf. "
            f"Headers found: {list(headers)}"
        )
    return resolved


def _detect_epsg_utm(
    lon: float,
    lat: float,
    fusos_path: str | None = FUSOS_UTM_PATH,
) -> int:
    """Auto-detect SIRGAS 2000 UTM EPSG for a given geographic point.

    Strategy:
    1. Spatial join of the point against FUSOS_UTM shapefile (MGRS grid).
       CODE format: '<zone><band>' e.g. '23K'. Bands C-M = UTM South, N-Z = UTM North.
       SIRGAS 2000 South EPSG = 31960 + zone, North EPSG = 31954 + zone.
    2. Fallback: longitude-based formula if shapefile is unavailable or join misses.
    3. Last resort: return EPSG_UTM (31983).
    """
    try:
        if fusos_path and Path(fusos_path).exists():
            import geopandas as gpd
            from shapely.geometry import Point

            fusos = gpd.read_file(fusos_path)
            # Align CRS: use fusos native CRS for the point to avoid sjoin warnings
            fusos_crs = fusos.crs or "EPSG:4326"
            pt = gpd.GeoDataFrame(geometry=[Point(lon, lat)], crs="EPSG:4674").to_crs(fusos_crs)
            joined = gpd.sjoin(pt, fusos[["CODE", "geometry"]], how="left", predicate="within")
            if not joined.empty:
                code = joined.iloc[0].get("CODE")
                if code and isinstance(code, str):
                    zone_match = re.search(r"\d+", str(code))
                    band_match = re.search(r"[A-Za-z]+", str(code))
                    if zone_match and band_match:
                        zone_num = int(zone_match.group())
                        is_south = band_match.group().upper() <= "M"
                        epsg = (31960 if is_south else 31954) + zone_num
                        if 31966 <= epsg <= 31992:  # valid SIRGAS 2000 UTM range
                            return epsg

        # Fallback: derive zone from longitude
        zone_num = int((lon + 180) / 6) + 1
        is_south = lat < 0
        return (31960 if is_south else 31954) + zone_num

    except Exception:
        return EPSG_UTM


def _centroid_from_gdf(gdf) -> tuple[float, float]:
    """Return (lon, lat) centroid of a GeoDataFrame in geographic coordinates."""
    bounds = gdf.to_crs(epsg=4674).total_bounds  # [minx, miny, maxx, maxy]
    return float((bounds[0] + bounds[2]) / 2), float((bounds[1] + bounds[3]) / 2)


def _load_snv_dataset(path: str, layer: str | None, epsg_src: int):
    try:
        import geopandas as gpd
        from pyproj import CRS
    except ImportError as exc:
        raise RuntimeError(
            "Missing geospatial dependencies. Install: python -m pip install geopandas shapely pyproj pyogrio"
        ) from exc

    read_kwargs = {}
    if layer:
        read_kwargs["layer"] = layer

    gdf = gpd.read_file(path, **read_kwargs)
    if gdf.empty:
        raise RuntimeError(f"SNV dataset is empty: {path}")
    if gdf.crs is None:
        gdf = gdf.set_crs(CRS.from_epsg(epsg_src))
    elif gdf.crs.to_epsg() != epsg_src:
        gdf = gdf.to_crs(epsg=epsg_src)

    mapping = _resolve_snv_fields(gdf.columns)
    gdf = gdf.copy()
    gdf["_br_norm"] = gdf[mapping["vl_br"]].apply(_normalize_br)
    gdf["_uf_norm"] = gdf[mapping["sg_uf"]].apply(_normalize_uf)
    gdf["_km_ini_norm"] = gdf[mapping["vl_km_inic"]].apply(_to_float)
    gdf["_km_fim_norm"] = gdf[mapping["vl_km_fina"]].apply(_to_float)
    return gdf, mapping


def _merge_geometry(parts):
    from shapely.geometry import LineString, MultiLineString
    from shapely.ops import linemerge

    if not parts:
        return None
    if len(parts) == 1:
        return parts[0]

    merged = linemerge(MultiLineString(parts))
    if isinstance(merged, (LineString, MultiLineString)):
        return merged
    if hasattr(merged, "geoms"):
        geoms = [geom for geom in merged.geoms if isinstance(geom, (LineString, MultiLineString))]
        if len(geoms) == 1:
            return geoms[0]
        line_geoms = [geom for geom in geoms if isinstance(geom, LineString)]
        if line_geoms:
            return MultiLineString(line_geoms)
    return MultiLineString(parts)


def _prepare_linestring(geometry):
    from shapely.geometry import LineString, MultiLineString
    from shapely.ops import linemerge

    if geometry is None or geometry.is_empty:
        return None
    if isinstance(geometry, LineString):
        return geometry
    if isinstance(geometry, MultiLineString):
        merged = linemerge(geometry)
        if isinstance(merged, LineString):
            return merged
        if hasattr(merged, "geoms"):
            return max(list(merged.geoms), key=lambda geom: geom.length)
    return geometry


def _extract_segment_geometry(
    gdf,
    mapping,
    br: object,
    uf: object,
    km_ini: float,
    km_fim: float,
    epsg_src: int,
    epsg_utm: int,
):
    try:
        import geopandas as gpd
        from shapely.ops import substring
    except ImportError as exc:
        raise RuntimeError(
            "Missing geospatial dependencies. Install: python -m pip install geopandas shapely pyproj pyogrio"
        ) from exc

    br_norm = _normalize_br(br)
    uf_norm = _normalize_uf(uf)
    requested_km_ini = km_ini
    requested_km_fim = km_fim
    reverse_requested = km_ini > km_fim
    km_req_ini = min(km_ini, km_fim)
    km_req_fim = max(km_ini, km_fim)
    subset = gdf[(gdf["_br_norm"] == br_norm) & (gdf["_uf_norm"] == uf_norm)].copy()
    if subset.empty:
        return {"error": f"No SNV feature found for BR-{br}/{uf_norm} in the supplied dataset."}

    overlap = subset[
        (subset["_km_ini_norm"] <= km_req_fim) & (subset["_km_fim_norm"] >= km_req_ini)
    ].sort_values("_km_ini_norm")
    if overlap.empty:
        return {
            "error": (
                f"No SNV feature for BR-{br_norm}/{uf_norm} covers km {requested_km_ini}-{requested_km_fim}. "
                f"Available extent in this dataset: km {subset['_km_ini_norm'].min():.3f}-"
                f"{subset['_km_fim_norm'].max():.3f}"
            )
        }

    parts = []
    rows = []
    for _, row in overlap.iterrows():
        row_km_ini = row["_km_ini_norm"]
        row_km_fim = row["_km_fim_norm"]
        if row_km_ini is None or row_km_fim is None or row_km_fim <= row_km_ini:
            continue

        geom = _prepare_linestring(row.geometry)
        if geom is None or geom.length <= 0:
            continue

        local_ini = max(km_req_ini, row_km_ini)
        local_fim = min(km_req_fim, row_km_fim)
        if local_fim <= local_ini:
            continue

        span = row_km_fim - row_km_ini
        frac_ini = max(0.0, min(1.0, (local_ini - row_km_ini) / span))
        frac_fim = max(0.0, min(1.0, (local_fim - row_km_ini) / span))
        segment = substring(geom, frac_ini * geom.length, frac_fim * geom.length)
        if segment is None or segment.is_empty:
            continue

        parts.append(segment)
        rows.append(
            {
                "geometry": segment,
                "br_uf": f"BR-{br_norm}/{uf_norm}",
                "km_ini_seg": round(local_fim if reverse_requested else local_ini, 3),
                "km_fim_seg": round(local_ini if reverse_requested else local_fim, 3),
                "km_ini_trecho": row_km_ini,
                "km_fim_trecho": row_km_fim,
                "vl_codigo": _clean_value(row.get(mapping.get("vl_codigo"))),
                "versao_snv": _clean_value(row.get(mapping.get("versao_snv"))),
            }
        )

    if not parts:
        return {"error": f"Geometry extraction returned no segment for BR-{br_norm}/{uf_norm} km {requested_km_ini}-{requested_km_fim}."}

    merged = _merge_geometry(parts)
    if reverse_requested and hasattr(merged, "reverse"):
        merged = merged.reverse()
    length_m = float(gpd.GeoSeries([merged], crs=f"EPSG:{epsg_src}").to_crs(epsg=epsg_utm).length.iloc[0])
    version_value = _clean_value(overlap.iloc[0].get(mapping.get("versao_snv")))
    return {
        "geometry": merged,
        "rows": rows,
        "meta": {
            "br_uf": f"BR-{br_norm}/{uf_norm}",
            "km_ini_req": requested_km_ini,
            "km_fim_req": requested_km_fim,
            "comprimento_m": round(length_m, 3),
            "comprimento_km": round(length_m / 1000.0, 6),
            "num_trechos": len(rows),
            "versao_snv": version_value or None,
            "epsg_geometria": epsg_src,
        },
    }


def _run_coord_lookup(gdf, mapping, lat: float, lon: float, raio_m: float, epsg_src: int, epsg_utm: int):
    try:
        import geopandas as gpd
        from shapely.geometry import Point
        from shapely.ops import nearest_points
    except ImportError as exc:
        raise RuntimeError(
            "Missing geospatial dependencies. Install: python -m pip install geopandas shapely pyproj pyogrio"
        ) from exc

    point_geo = gpd.GeoSeries([Point(lon, lat)], crs=f"EPSG:{epsg_src}")
    point_utm = point_geo.to_crs(epsg=epsg_utm).iloc[0]
    gdf_utm = gdf.to_crs(epsg=epsg_utm)
    search = gdf_utm[gdf_utm.geometry.intersects(point_utm.buffer(raio_m))]
    if search.empty:
        return None

    best_index = None
    best_distance = math.inf
    best_geom = None
    for idx, row in search.iterrows():
        distance = point_utm.distance(row.geometry)
        if distance < best_distance:
            best_index = idx
            best_distance = distance
            best_geom = row.geometry

    src_row = gdf.loc[best_index]
    geom = _prepare_linestring(best_geom)
    near_point = nearest_points(point_utm, geom)[1]
    fraction = 0.0 if geom.length <= 0 else max(0.0, min(1.0, geom.project(near_point) / geom.length))

    km_ini = _to_float(src_row[mapping["vl_km_inic"]])
    km_fim = _to_float(src_row[mapping["vl_km_fina"]])
    if km_ini is None or km_fim is None:
        raise RuntimeError("Selected SNV feature has invalid km fields.")

    return {
        "br_uf": f"BR-{_clean_value(src_row[mapping['vl_br']])}/{_clean_value(src_row[mapping['sg_uf']]).upper()}",
        "km": round(km_ini + fraction * (km_fim - km_ini), 3),
        "dist_lat_m": round(best_distance, 3) if best_distance > 10 else None,
        "vl_codigo": _clean_value(src_row.get(mapping.get("vl_codigo"))) or None,
        "versao_snv": _clean_value(src_row.get(mapping.get("versao_snv"))) or None,
        "km_ini_attr": km_ini,
        "km_fim_attr": km_fim,
        "frac_trecho": round(fraction, 6),
        "epsg_medicoes": epsg_utm,
    }


def _segment_output_result(segment_result, formato: str, output_path: str | None, epsg_src: int):
    meta = dict(segment_result["meta"])
    geometry = segment_result["geometry"]
    rows = segment_result["rows"]

    if formato == "wkt":
        return {**meta, "wkt": geometry.wkt}

    if formato in {"shp", "gpkg"}:
        if not output_path:
            raise RuntimeError(f"--output is required when --formato={formato}")
        try:
            import geopandas as gpd
        except ImportError as exc:
            raise RuntimeError(
                "Missing geospatial dependencies. Install: python -m pip install geopandas shapely pyproj pyogrio"
            ) from exc

        gdf_out = gpd.GeoDataFrame(rows, geometry="geometry", crs=f"EPSG:{epsg_src}")
        if formato == "shp":
            gdf_out.to_file(output_path, driver="ESRI Shapefile", encoding="utf-8")
        else:
            layer_name = Path(output_path).stem
            gdf_out.to_file(output_path, driver="GPKG", layer=layer_name)
        return {**meta, "arquivo_gerado": output_path, "formato": formato, "num_feicoes": len(gdf_out)}

    features = []
    for row in rows:
        props = {key: value for key, value in row.items() if key != "geometry"}
        features.append(
            {
                "type": "Feature",
                "geometry": row["geometry"].__geo_interface__,
                "properties": props,
            }
        )

    return {
        **meta,
        "geojson": {"type": "FeatureCollection", "features": features},
        "geojson_merged": {
            "type": "Feature",
            "geometry": geometry.__geo_interface__,
            "properties": {
                "br_uf": meta["br_uf"],
                "km_ini": meta["km_ini_req"],
                "km_fim": meta["km_fim_req"],
                "comprimento_m": meta["comprimento_m"],
            },
        },
    }


def _default_output_path(input_path: str, output_path: str | None) -> str:
    if output_path:
        return output_path
    source = Path(input_path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        return str(source.with_name(f"{source.stem}_snv_wkt.csv"))
    return str(source.with_name(f"{source.stem}_snv_wkt.xlsx"))


def _read_csv_table(path: str):
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        except csv.Error:
            class Fallback(csv.excel):
                delimiter = ";"
            dialect = Fallback
        reader = csv.DictReader(handle, dialect=dialect)
        rows = [dict(row) for row in reader]
        return {
            "kind": "csv",
            "headers": list(reader.fieldnames or []),
            "rows": rows,
            "dialect": dialect,
        }


def _write_csv_table(path: str, headers, rows, dialect):
    with open(path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, dialect=dialect)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _read_xlsx_table(path: str, sheet_name: str | None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xlsx requires openpyxl. Install: python -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(path)
    actual_sheet = _pick_sheet_name(workbook.sheetnames, sheet_name)
    worksheet = workbook[actual_sheet]
    headers = _trim_headers([cell.value for cell in worksheet[1]])
    if not headers:
        raise RuntimeError(f"Worksheet '{actual_sheet}' has no header row.")

    rows = []
    for row_index in range(2, worksheet.max_row + 1):
        row_data = {}
        for col_index, header in enumerate(headers, start=1):
            row_data[header] = _clean_value(worksheet.cell(row=row_index, column=col_index).value)
        if any(value != "" for value in row_data.values()):
            rows.append(row_data)

    return {
        "kind": "xlsx",
        "headers": headers,
        "rows": rows,
        "sheet_name": actual_sheet,
        "workbook": workbook,
    }


def _read_xls_table(path: str, sheet_name: str | None):
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xls requires xlrd. Install: python -m pip install xlrd"
        ) from exc

    workbook = xlrd.open_workbook(path)
    actual_sheet = _pick_sheet_name(workbook.sheet_names(), sheet_name)
    worksheet = workbook.sheet_by_name(actual_sheet)
    if worksheet.nrows == 0:
        raise RuntimeError(f"Worksheet '{actual_sheet}' has no rows.")

    headers = _trim_headers(worksheet.row_values(0))
    if not headers:
        raise RuntimeError(f"Worksheet '{actual_sheet}' has no header row.")

    rows = []
    for row_index in range(1, worksheet.nrows):
        values = worksheet.row_values(row_index)
        row_data = {}
        for col_index, header in enumerate(headers):
            cell_value = values[col_index] if col_index < len(values) else ""
            row_data[header] = _clean_value(cell_value)
        if any(value != "" for value in row_data.values()):
            rows.append(row_data)

    return {
        "kind": "xls",
        "headers": headers,
        "rows": rows,
        "sheet_name": actual_sheet,
    }


def _write_new_xlsx(path: str, headers, rows, sheet_name: str):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "Writing .xlsx requires openpyxl. Install: python -m pip install openpyxl"
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or DEFAULT_TABLE_SHEET
    for col_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_index, value=header)
    for row_index, row in enumerate(rows, start=2):
        for col_index, header in enumerate(headers, start=1):
            worksheet.cell(row=row_index, column=col_index, value=row.get(header, ""))
    workbook.save(path)


def _write_preserved_xlsx(path: str, table, managed_headers, rows):
    workbook = table["workbook"]
    worksheet = workbook[table["sheet_name"]]

    header_positions = {}
    for col_index in range(1, worksheet.max_column + 1):
        header = _clean_value(worksheet.cell(row=1, column=col_index).value)
        if header:
            header_positions[header] = col_index

    next_column = max(header_positions.values(), default=0) + 1
    for header in managed_headers:
        if header not in header_positions:
            header_positions[header] = next_column
            worksheet.cell(row=1, column=next_column, value=header)
            next_column += 1

    for row_index, row in enumerate(rows, start=2):
        for header in managed_headers:
            worksheet.cell(row=row_index, column=header_positions[header], value=row.get(header, ""))

    workbook.save(path)


def _read_input_table(path: str, sheet_name: str | None):
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return _read_csv_table(path)
    if suffix == ".xlsx":
        return _read_xlsx_table(path, sheet_name)
    if suffix == ".xls":
        return _read_xls_table(path, sheet_name)
    raise RuntimeError(f"Unsupported table format: {suffix}. Use .csv, .xlsx, or .xls.")


def _write_output_table(path: str, table, headers, rows, managed_headers):
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        dialect = table.get("dialect", csv.excel)
        _write_csv_table(path, headers, rows, dialect)
        return
    if suffix == ".xlsx":
        if table["kind"] == "xlsx":
            _write_preserved_xlsx(path, table, managed_headers, rows)
        else:
            _write_new_xlsx(path, headers, rows, table.get("sheet_name") or DEFAULT_TABLE_SHEET)
        return
    raise RuntimeError(f"Unsupported output format: {suffix}. Use .csv or .xlsx.")


def _run_table_mode(
    gdf,
    mapping,
    input_path: str,
    output_path: str | None,
    sheet_name: str | None,
    wkt_column: str,
    error_column: str,
    uf_default: str,
    epsg_src: int,
    epsg_utm: int,
):
    table = _read_input_table(input_path, sheet_name)
    output_path = _default_output_path(input_path, output_path)
    field_map = _resolve_table_fields(table["headers"])
    headers = list(table["headers"])
    managed_headers = []

    if wkt_column not in headers:
        headers.append(wkt_column)
    managed_headers.append(wkt_column)

    output_rows = []
    has_errors = False
    processed = 0

    for row_number, row in enumerate(table["rows"], start=2):
        row_out = dict(row)
        br_value = row.get(field_map["br"])
        uf_value = row.get(field_map["uf"]) if field_map["uf"] else uf_default
        km_ini = _to_float(row.get(field_map["kmi"]))
        km_fim = _to_float(row.get(field_map["kmf"]))

        if not _normalize_br(br_value):
            row_out[wkt_column] = ""
            row_out[error_column] = f"Row {row_number}: invalid BR value"
            has_errors = True
            output_rows.append(row_out)
            continue
        if km_ini is None or km_fim is None:
            row_out[wkt_column] = ""
            row_out[error_column] = f"Row {row_number}: invalid kmi/kmf values"
            has_errors = True
            output_rows.append(row_out)
            continue
        if km_fim == km_ini:
            row_out[wkt_column] = ""
            row_out[error_column] = f"Row {row_number}: kmf must be different from kmi"
            has_errors = True
            output_rows.append(row_out)
            continue

        result = _extract_segment_geometry(
            gdf,
            mapping,
            br_value,
            _normalize_uf(uf_value, uf_default),
            km_ini,
            km_fim,
            epsg_src,
            epsg_utm,
        )
        if "error" in result:
            row_out[wkt_column] = ""
            row_out[error_column] = f"Row {row_number}: {result['error']}"
            has_errors = True
        else:
            row_out[wkt_column] = result["geometry"].wkt
            row_out[error_column] = ""
            processed += 1

        output_rows.append(row_out)

    if has_errors or error_column in headers:
        if error_column not in headers:
            headers.append(error_column)
        managed_headers.append(error_column)

    _write_output_table(output_path, table, headers, output_rows, managed_headers)
    return {
        "input": input_path,
        "output": output_path,
        "sheet_name": table.get("sheet_name"),
        "rows_total": len(output_rows),
        "rows_ok": processed,
        "rows_error": len(output_rows) - processed,
        "wkt_column": wkt_column,
        "error_column": error_column if has_errors else None,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="SNV tools for point lookup, segment extraction, and table enrichment.")
    parser.add_argument("snv_path", help="Path to the SNV dataset (.shp or .gpkg)")
    parser.add_argument("--modo", choices=["coord", "segmento", "tabela"], default="coord")

    parser.add_argument("--lat", type=float, help="[coord] Latitude in decimal degrees")
    parser.add_argument("--lon", type=float, help="[coord] Longitude in decimal degrees")
    parser.add_argument("--raio", type=float, default=250.0, help="[coord] Search radius in meters")

    parser.add_argument("--br", type=str, help="[segmento] BR number")
    parser.add_argument("--uf", type=str, default="MG", help="[segmento] State abbreviation. Default: MG")
    parser.add_argument("--km-ini", dest="km_ini", type=float, help="[segmento] Initial kilometer")
    parser.add_argument("--km-fim", dest="km_fim", type=float, help="[segmento] Final kilometer")
    parser.add_argument("--formato", choices=["json", "wkt", "shp", "gpkg"], default="json")
    parser.add_argument("--output", type=str, help="[segmento|tabela] Output path")

    parser.add_argument("--input", type=str, help="[tabela] Input table path (.csv, .xlsx, .xls)")
    parser.add_argument("--sheet", type=str, default=DEFAULT_TABLE_SHEET, help="[tabela] Worksheet name. Default: Export_ICM")
    parser.add_argument("--wkt-column", type=str, default="geometry_wkt", help="[tabela] Name of the WKT output column")
    parser.add_argument("--error-column", type=str, default="snv_error", help="[tabela] Name of the row error column")
    parser.add_argument("--uf-default", type=str, default="MG", help="[tabela] Default UF when the table column is blank or absent")

    parser.add_argument("--snv-layer", type=str, default=None, help="Layer name when the SNV input is a GeoPackage")
    parser.add_argument("--epsg-src", type=int, default=EPSG_GEO, help="Input SNV EPSG. Default: 4674")
    parser.add_argument("--epsg-utm", type=int, default=0, help="Metric EPSG for calculations. Default: auto-detected from dataset centroid via FUSOS_UTM shapefile")
    parser.add_argument("--fusos-path", type=str, default=FUSOS_UTM_PATH, help="Path to FUSOS_UTM shapefile for UTM zone auto-detection")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    try:
        gdf, mapping = _load_snv_dataset(args.snv_path, args.snv_layer, args.epsg_src)

        if args.epsg_utm != 0:
            epsg_utm = args.epsg_utm
        elif args.modo == "coord" and args.lat is not None and args.lon is not None:
            epsg_utm = _detect_epsg_utm(args.lon, args.lat, args.fusos_path)
        else:
            cx, cy = _centroid_from_gdf(gdf)
            epsg_utm = _detect_epsg_utm(cx, cy, args.fusos_path)

        if args.modo == "coord":
            if args.lat is None or args.lon is None:
                parser.error("Mode 'coord' requires --lat and --lon")
            result = _run_coord_lookup(gdf, mapping, args.lat, args.lon, args.raio, args.epsg_src, epsg_utm)
            if result is None:
                raise RuntimeError(
                    f"No SNV feature found within {args.raio} m of lat={args.lat}, lon={args.lon}. "
                    "Try a larger radius with --raio."
                )

        elif args.modo == "segmento":
            missing = [flag for flag, value in [("--br", args.br), ("--km-ini", args.km_ini), ("--km-fim", args.km_fim)] if value is None]
            if missing:
                parser.error(f"Mode 'segmento' requires: {', '.join(missing)}")
            if args.km_fim == args.km_ini:
                parser.error("--km-fim must be different from --km-ini")
            segment_result = _extract_segment_geometry(
                gdf,
                mapping,
                args.br,
                args.uf,
                args.km_ini,
                args.km_fim,
                args.epsg_src,
                epsg_utm,
            )
            if "error" in segment_result:
                raise RuntimeError(segment_result["error"])
            result = _segment_output_result(segment_result, args.formato, args.output, args.epsg_src)

        else:
            if not args.input:
                parser.error("Mode 'tabela' requires --input")
            result = _run_table_mode(
                gdf,
                mapping,
                args.input,
                args.output,
                args.sheet,
                args.wkt_column,
                args.error_column,
                args.uf_default,
                args.epsg_src,
                epsg_utm,
            )

        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    except SystemExit:
        raise
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False, indent=2))
        return 1


if __name__ == "__main__":
    sys.exit(main())
